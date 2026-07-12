#!/usr/bin/env python3
"""
excel-incell-translate :: embedder
将译文嵌入【源单元格内】（原文在上、译文在下，换行分隔），不新增列、不修改原文。

两种模式：
  A. 恢复模式（--recover --backup）：文件列数翻倍（上次误把译文放到了右侧独立列），
     复用右侧列里已算好的干净正文译文，并删除右列；表头用外部字典。
  B. 翻译模式（默认，需 --dict）：提供外部字典 JSON，对每个文本单元格现译现嵌。

幂等 / 增量（最佳实践）：
  - 重跑时已嵌好译文的单元格（原文\\n译文，两行且末行语言相反）会被 is_already_translated()
    识别并跳过，绝不重复翻译 -> 已翻译的文件可安全重复运行，只补未译部分。

通用规则：
  - 公式 / 数字 / 空白 / 纯标点 -> 跳过，保持原样
  - 无法翻译的 token（用例编号、缩写、运营商/人名）-> 跳过，不乱写
  - 写入 source + "\\n" + translation，开启 wrap_text，保留原对齐
  - 保存前设 fullCalcOnLoad=True（公式打开即重算）
  - 验证：列数复原 / 原文保留 / 公式串一致

用法：
  # 翻译模式（需外部字典；表头与正文共用同一份 {source: translation}）
  python embed_incell.py <workbook.xlsx> --dict dict.json [--backup clean.xlsx]
  # 恢复模式（从右侧错位列回收，必须带 --backup 以安全判定列数）
  python embed_incell.py <workbook.xlsx> --recover --backup clean.xlsx
"""
import argparse, json, re, sys
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

cjk = re.compile(r'[\u4e00-\u9fff]')
lat = re.compile(r'[A-Za-z]')

# 判断性 / 状态结论字符（通用测试结论词）：一律不翻译，保持原样。
# 这些词是测试结论（通过 / 失败 / 不适用 …），译或不译对任何人含义一致，
# 且常被自动化解析，追加译文反而有害。本规则优先级高于字典——
# 即便字典里写了 PASS→通过，脚本也强制跳过，绝不修改这类单元格。
STATUS_SKIP = frozenset({
    'PASS', 'FAIL', 'FAILED', 'PASS/FAIL',
    'N/A', 'NA',
    'OK', 'NG',
    'WIP', 'TBD', 'PENDING', 'BLOCKED', 'SKIP', 'SKIPPED', 'DONE',
    'YES', 'NO', 'TRUE', 'FALSE',
    '通过', '失败', '不适用',
})


def is_status_token(text):
    """判断性 / 状态结论字符（整格精确匹配，忽略大小写）。"""
    if not isinstance(text, str):
        return False
    s = text.strip()
    return s.upper() in STATUS_SKIP or s in STATUS_SKIP


def has_cjk(s):
    return bool(cjk.search(s or ''))


def has_lat(s):
    return bool(lat.search(s or ''))


def lang_of(s):
    s = (s or '').strip()
    c = has_cjk(s)
    l = has_lat(s)
    if c and not l:
        return 'cn'
    if l and not c:
        return 'en'
    if c and l:
        return 'mix'
    return 'none'


def direction_of(val):
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    if has_cjk(s):
        return 'zh2en'      # 含中文 -> 译英文
    if has_lat(s):
        return 'en2zh'      # 纯英文 -> 译中文
    return None


def is_already_translated(value):
    """幂等判定（最佳实践 #1：幂等优先）。
    单元格已是 '原文\\n译文' 形式（至少两行，且首行与末行语言相反）即视为已译，跳过。
    对已经生成的双语文件天然生效，避免重跑时重复翻译或叠加译文（如 'PASS\\n通过\\n<乱译>'）。
    注意：极少情况下用户自己手写的中英两行也可能被误判为已译而跳过（风险低，已与用户确认）。
    """
    if not isinstance(value, str):
        return False
    lines = value.split('\n')
    if len(lines) < 2:
        return False
    first = lang_of(lines[0])
    last = lang_of(lines[-1])
    if first in ('cn', 'mix') and last == 'en':
        return True
    if first == 'en' and last in ('cn', 'mix'):
        return True
    return False


def translate(text, dictionary):
    """从统一外部字典查译文；无译文或译文等于原文返回 None（宁漏不错，不乱写）。
    判断性 / 状态结论字符（PASS/FAIL/N/A…）强制跳过，优先级高于字典。
    """
    if not isinstance(text, str) or not text.strip():
        return None
    if is_status_token(text):
        return None
    t = dictionary.get(text.strip())
    if t and t != text.strip():
        return t
    return None


def embed(path, dictionary, backup=None, recover=False):
    wb = openpyxl.load_workbook(path)                      # 保留样式
    wb_v = openpyxl.load_workbook(path, data_only=True)    # 取值
    wb_b = openpyxl.load_workbook(backup) if backup else None

    total = 0
    skipped = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        ws_v = wb_v[sn]
        mm = ws.max_column
        rows = ws.max_row
        if mm == 0 or rows == 0:
            continue
        # 真实源列数：有备份用备份；恢复模式用一半；否则用当前列数
        if wb_b:
            bm = wb_b[sn].max_column
        elif recover:
            bm = mm // 2 if mm % 2 == 0 else mm
        else:
            bm = mm
        doubled = recover and (mm == 2 * bm)
        sheet_n = 0
        sheet_skip = 0
        for r in range(1, rows + 1):
            for c in range(1, bm + 1):
                src = ws.cell(row=r, column=c)
                if isinstance(src, MergedCell):
                    continue
                orig = src.value
                # 公式：跳过（保持原样）
                if isinstance(orig, str) and orig.startswith('='):
                    continue
                # 幂等：已译则跳过（重跑只补未译）
                if is_already_translated(orig):
                    sheet_skip += 1
                    continue
                # ---- 恢复模式：复用右侧干净正文译文 ----
                if doubled:
                    if r == 1:
                        tr = translate(orig, dictionary) if isinstance(orig, str) else None
                        if tr:
                            src.value = f"{orig}\n{tr}"
                            _wrap(src)
                            sheet_n += 1
                        continue
                    old_tr = ws.cell(row=r, column=bm + c).value
                    if isinstance(orig, str) and isinstance(old_tr, str) \
                            and old_tr.strip() and old_tr != orig.strip() \
                            and not is_status_token(orig):
                        src.value = f"{orig}\n{old_tr}"
                        _wrap(src)
                        sheet_n += 1
                    continue
                # ---- 翻译模式：现译现嵌（表头与正文共用同一外部字典）----
                if isinstance(orig, str):
                    tr = translate(orig.strip(), dictionary)
                    if tr:
                        src.value = f"{orig}\n{tr}"
                        _wrap(src)
                        sheet_n += 1
                # 非字符串（数字等）或字典无对应译文：跳过
        if doubled and mm > bm:
            ws.delete_cols(bm + 1, mm - bm)   # 删除错位右列
        total += sheet_n
        skipped += sheet_skip
        print(f"  {sn:24s} 嵌入 {sheet_n:4d}  跳过已译 {sheet_skip:4d}"
              + ("  删除右列" if doubled else ""))
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(path)
    print(f"\n已保存。共嵌入 {total} 处，跳过已译 {skipped} 处。")
    verify(path, backup)


def _wrap(cell):
    al = cell.alignment
    cell.alignment = Alignment(horizontal=al.horizontal,
                               vertical=al.vertical, wrap_text=True)


def verify(path, backup):
    print("\n验证：列数复原 / 原文保留 / 公式不变")
    wb_nf = openpyxl.load_workbook(path)
    wb_nv = openpyxl.load_workbook(path, data_only=True)
    wb_bk = openpyxl.load_workbook(backup) if backup else None
    col_ok, fdiff, preserved, broken = True, 0, 0, 0
    if not wb_bk:
        print("  （未提供 --backup，跳过列数/原文/公式的逐格比对；"
              "建议翻译前保留干净副本用于校验）")
    for sn in wb_nf.sheetnames:
        bm = (wb_bk[sn].max_column if wb_bk
              else wb_nf[sn].max_column)
        nm = wb_nf[sn].max_column
        if wb_bk and bm != nm:
            col_ok = False
            print(f"  ❌ 列数未复原 {sn}: 期望 {bm} 实际 {nm}")
        wsf = wb_nf[sn]
        wbk_sn = wb_bk[sn] if wb_bk else None
        rows = wsf.max_row
        for r in range(1, rows + 1):
            for c in range(1, min(bm, nm) + 1):
                ofv = wbk_sn.cell(r, c).value if wbk_sn else None
                nfv = wsf.cell(r, c).value
                if isinstance(ofv, str) and ofv.startswith('='):
                    if ofv != nfv:
                        fdiff += 1
                    continue
                if isinstance(ofv, str) and ofv and isinstance(nfv, str):
                    if nfv.startswith(ofv) or ofv in nfv:
                        preserved += 1
                    else:
                        broken += 1
                        if broken <= 5:
                            print(f"  ⚠ 原文疑似被改 {sn} r{r}c{c}: "
                                  f"原={ofv[:25]!r} 新={nfv[:25]!r}")
    print(f"列数复原: {'✅' if col_ok else '✅(无备份未比对)' if not wb_bk else '❌'} "
          f"| 公式差异: {fdiff} | 原文保留: {preserved} 疑似破坏: {broken}")


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('workbook')
    ap.add_argument('--dict', default=None,
                    help='外部翻译字典 JSON: {"source": "translation"}，表头与正文共用')
    ap.add_argument('--backup', default=None,
                    help='干净源副本，用于列数复原核查与恢复模式列数判定')
    ap.add_argument('--recover', action='store_true',
                    help='恢复模式：从右侧错位列回收译文（必须同时提供 --backup）')
    a = ap.parse_args()
    dictionary = {}
    if a.dict:
        with open(a.dict, encoding='utf-8') as f:
            dictionary = json.load(f)
    if a.recover and not a.backup:
        sys.exit("❌ 恢复模式需要 --backup 干净副本以安全判定列数，已终止。")
    if not a.recover and not dictionary:
        sys.exit("❌ 翻译模式需要 --dict 外部字典；或使用 --recover 恢复模式。")
    embed(a.workbook, dictionary, a.backup, a.recover)
